#!/usr/bin/env python3
"""
CEFR-J Wordlist Ver1.6 Excel to CSV Converter
A1-B2レベルの語彙をCSV形式に変換
"""

import openpyxl
import csv
import json
import sys
from pathlib import Path

def convert_excel_to_csv(excel_path: str, output_csv: str):
    """
    CEFR-J WordlistのExcelファイルをCSVに変換
    
    Args:
        excel_path: 入力Excelファイルパス
        output_csv: 出力CSVファイルパス
    """
    print(f"📂 Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # シート名を確認
    print(f"📋 Available sheets: {wb.sheetnames}")
    
    # A1-B2の各レベルシートを処理
    levels = ['A1', 'A2', 'B1', 'B2']
    all_words = []
    
    for level in levels:
        # _sep版（分割版）を優先的に使用
        sheet_name = f"{level}_sep"
        if sheet_name not in wb.sheetnames:
            sheet_name = level
        
        if sheet_name not in wb.sheetnames:
            print(f"⚠️ Sheet {sheet_name} not found, skipping...")
            continue
        
        print(f"📖 Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        # ヘッダー行を取得（通常は1行目）
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        print(f"   Headers: {headers[:5]}...")  # 最初の5列を表示
        
        # データ行を処理
        row_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 最初の列が空ならスキップ
                continue
            
            # 単語データを抽出
            word_data = {
                'word': str(row[0]).strip() if row[0] else '',
                'cefr_level': level,
                'pos': str(row[1]).strip() if len(row) > 1 and row[1] else 'unknown',
            }
            
            # 空の単語はスキップ
            if not word_data['word'] or word_data['word'] == 'None':
                continue
            
            all_words.append(word_data)
            row_count += 1
        
        print(f"   ✅ Processed {row_count} words from {level}")
    
    # CSVに書き込み
    print(f"\n💾 Writing to CSV: {output_csv}")
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['word', 'pos', 'cefr_level'])
        writer.writeheader()
        writer.writerows(all_words)
    
    print(f"✅ Successfully wrote {len(all_words)} words to CSV")
    
    # 統計情報を表示
    level_counts = {}
    for word in all_words:
        level = word['cefr_level']
        level_counts[level] = level_counts.get(level, 0) + 1
    
    print("\n📊 Statistics by CEFR Level:")
    for level in ['A1', 'A2', 'B1', 'B2']:
        count = level_counts.get(level, 0)
        print(f"   {level}: {count:,} words")
    
    return len(all_words)

def main():
    # パス設定
    base_dir = Path(__file__).parent.parent
    excel_path = base_dir / "data" / "vocabulary" / "cefrj_wordlist_v16.xlsx"
    output_csv = base_dir / "data" / "vocabulary" / "cefrj_wordlist_parsed.csv"
    
    if not excel_path.exists():
        print(f"❌ Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    try:
        total_words = convert_excel_to_csv(str(excel_path), str(output_csv))
        print(f"\n🎉 Conversion completed successfully!")
        print(f"📁 Output file: {output_csv}")
        print(f"📊 Total words: {total_words:,}")
        return 0
    except Exception as e:
        print(f"❌ Error during conversion: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
